"""
Advanced Analytics Export Module

Provides comprehensive export functionality for portfolio analytics results including:
- CSV exports for quantitative data
- JSON exports for structured results
- Excel exports with multiple worksheets
- PDF reports with professional formatting
"""

import numpy as np
import pandas as pd
import json
import logging
from typing import Dict, List, Any, Optional
from datetime import datetime
from pathlib import Path
import warnings

# Optional dependencies - handle gracefully
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.charts import BarChart, LineChart, Reference
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

logger = logging.getLogger(__name__)

# Suppress openpyxl warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


class AnalyticsExporter:
    """
    Exports advanced analytics results to multiple formats.
    
    Supports export to:
    - CSV files for quantitative analysis
    - JSON for API integration and data interchange
    - Excel workbooks with formatted sheets and charts
    - PDF reports with professional formatting
    """
    
    def __init__(self, output_dir: str = "/mnt/optimizer_share/output"):
        """
        Initialize analytics exporter
        
        Args:
            output_dir: Directory to save exported files
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Setup styles if available
        if HAS_REPORTLAB:
            self.styles = getSampleStyleSheet()
            self._setup_custom_styles()
        else:
            self.styles = None
            self.custom_styles = {}
        
    def _setup_custom_styles(self):
        """Setup custom styles for PDF reports"""
        self.custom_styles = {
            'CustomTitle': ParagraphStyle(
                'CustomTitle',
                parent=self.styles['Title'],
                fontSize=18,
                spaceAfter=30,
                textColor=colors.darkblue,
                alignment=1  # Center alignment
            ),
            'CustomHeading': ParagraphStyle(
                'CustomHeading',
                parent=self.styles['Heading1'],
                fontSize=14,
                spaceAfter=12,
                textColor=colors.darkblue
            ),
            'CustomSubHeading': ParagraphStyle(
                'CustomSubHeading',
                parent=self.styles['Heading2'],
                fontSize=12,
                spaceAfter=8,
                textColor=colors.darkgreen
            )
        }
    
    def export_to_csv(self, analytics_results: Dict[str, Any],
                     filename_prefix: str = "analytics_export") -> Dict[str, str]:
        """
        Export analytics results to CSV files
        
        Args:
            analytics_results: Combined analytics results
            filename_prefix: Prefix for generated filenames
            
        Returns:
            Dictionary mapping data types to CSV file paths
        """
        logger.info("📄 Exporting analytics results to CSV files")
        
        csv_files = {}
        
        try:
            # Portfolio summary CSV
            if 'attribution' in analytics_results:
                portfolio_summary = analytics_results['attribution'].get('portfolio_summary', {})
                if portfolio_summary:
                    summary_df = pd.DataFrame([portfolio_summary])
                    summary_path = self.output_dir / f"{filename_prefix}_portfolio_summary_{self.timestamp}.csv"
                    summary_df.to_csv(summary_path, index=False)
                    csv_files['portfolio_summary'] = str(summary_path)
            
            # Strategy contributions CSV
            if 'attribution' in analytics_results:
                strategy_contrib = analytics_results['attribution'].get('strategy_contribution', {})
                if strategy_contrib:
                    contrib_data = []
                    for strategy_name, contrib_info in strategy_contrib.items():
                        contrib_data.append({
                            'strategy_name': strategy_name,
                            'total_contribution': contrib_info.get('total_contribution', 0),
                            'avg_daily_contribution': contrib_info.get('avg_daily_contribution', 0),
                            'volatility_contribution': contrib_info.get('volatility_contribution', 0),
                            'weight': contrib_info.get('weight', 0)
                        })
                    contrib_df = pd.DataFrame(contrib_data)
                    contrib_path = self.output_dir / f"{filename_prefix}_strategy_contributions_{self.timestamp}.csv"
                    contrib_df.to_csv(contrib_path, index=False)
                    csv_files['strategy_contributions'] = str(contrib_path)
            
            # Stop Loss attribution CSV
            if 'attribution' in analytics_results:
                sl_attribution = analytics_results['attribution'].get('stop_loss_attribution', {})
                if sl_attribution:
                    sl_data = []
                    for sl_level, sl_info in sl_attribution.items():
                        sl_data.append({
                            'stop_loss_level': sl_level,
                            'total_return': sl_info.get('total_return', 0),
                            'avg_daily_return': sl_info.get('avg_daily_return', 0),
                            'volatility': sl_info.get('volatility', 0),
                            'weight_in_portfolio': sl_info.get('weight_in_portfolio', 0),
                            'num_strategies': sl_info.get('num_strategies', 0)
                        })
                    sl_df = pd.DataFrame(sl_data)
                    sl_path = self.output_dir / f"{filename_prefix}_stop_loss_attribution_{self.timestamp}.csv"
                    sl_df.to_csv(sl_path, index=False)
                    csv_files['stop_loss_attribution'] = str(sl_path)
            
            # Risk metrics CSV
            if 'risk_metrics' in analytics_results:
                risk_data = analytics_results['risk_metrics']
                
                # VaR analysis
                var_analysis = risk_data.get('var_analysis', {})
                if var_analysis:
                    var_df = pd.DataFrame([var_analysis])
                    var_path = self.output_dir / f"{filename_prefix}_var_analysis_{self.timestamp}.csv"
                    var_df.to_csv(var_path, index=False)
                    csv_files['var_analysis'] = str(var_path)
                
                # Drawdown analysis
                drawdown_analysis = risk_data.get('drawdown_analysis', {})
                if drawdown_analysis and 'daily_drawdowns' in drawdown_analysis:
                    drawdown_df = pd.DataFrame({
                        'day': range(len(drawdown_analysis['daily_drawdowns'])),
                        'drawdown': drawdown_analysis['daily_drawdowns']
                    })
                    drawdown_path = self.output_dir / f"{filename_prefix}_drawdown_analysis_{self.timestamp}.csv"
                    drawdown_df.to_csv(drawdown_path, index=False)
                    csv_files['drawdown_analysis'] = str(drawdown_path)
            
            # Sensitivity analysis CSV
            if 'sensitivity' in analytics_results:
                sensitivity_data = analytics_results['sensitivity']
                
                # Correlation penalty sensitivity
                corr_penalty_sens = sensitivity_data.get('correlation_penalty_sensitivity', {})
                if corr_penalty_sens and 'penalty_results' in corr_penalty_sens:
                    penalty_data = []
                    for key, result in corr_penalty_sens['penalty_results'].items():
                        if 'error' not in result:
                            penalty_data.append({
                                'penalty_weight': result.get('penalty_weight', 0),
                                'fitness': result.get('fitness', 0),
                                'portfolio_size': result.get('portfolio_size', 0)
                            })
                    if penalty_data:
                        penalty_df = pd.DataFrame(penalty_data)
                        penalty_path = self.output_dir / f"{filename_prefix}_correlation_penalty_sensitivity_{self.timestamp}.csv"
                        penalty_df.to_csv(penalty_path, index=False)
                        csv_files['correlation_penalty_sensitivity'] = str(penalty_path)
            
            logger.info(f"✅ Exported {len(csv_files)} CSV files")
            
        except Exception as e:
            logger.error(f"❌ Error exporting CSV files: {e}")
        
        return csv_files
    
    def export_to_json(self, analytics_results: Dict[str, Any],
                      filename: str = "analytics_export.json") -> str:
        """
        Export analytics results to JSON format
        
        Args:
            analytics_results: Combined analytics results
            filename: Output filename
            
        Returns:
            Path to exported JSON file
        """
        logger.info("📄 Exporting analytics results to JSON format")
        
        try:
            # Clean results for JSON serialization
            json_results = self._prepare_json_data(analytics_results)
            
            # Add metadata
            json_results['export_metadata'] = {
                'export_timestamp': datetime.now().isoformat(),
                'data_source': 'Heavy Optimizer Platform',
                'strategies_analyzed': len(analytics_results.get('strategy_names', [])),
                'trading_days': analytics_results.get('trading_days', 82),
                'analysis_components': list(json_results.keys())
            }
            
            # Save JSON file
            json_path = self.output_dir / f"{filename.replace('.json', '')}_{self.timestamp}.json"
            with open(json_path, 'w') as f:
                json.dump(json_results, f, indent=2, default=self._json_serializer)
            
            logger.info(f"✅ JSON export saved to {json_path}")
            return str(json_path)
            
        except Exception as e:
            logger.error(f"❌ Error exporting JSON: {e}")
            return ""
    
    def _prepare_json_data(self, data: Any) -> Any:
        """Prepare data for JSON serialization by converting numpy arrays and other non-serializable types"""
        if isinstance(data, dict):
            return {key: self._prepare_json_data(value) for key, value in data.items()}
        elif isinstance(data, list):
            return [self._prepare_json_data(item) for item in data]
        elif isinstance(data, np.ndarray):
            return data.tolist()
        elif isinstance(data, (np.integer, np.floating)):
            return float(data)
        elif isinstance(data, np.bool_):
            return bool(data)
        elif pd.isna(data):
            return None
        else:
            return data
    
    def _json_serializer(self, obj):
        """Custom JSON serializer for complex objects"""
        if isinstance(obj, (np.ndarray, np.generic)):
            return obj.tolist()
        elif isinstance(obj, pd.Timestamp):
            return obj.isoformat()
        elif hasattr(obj, 'isoformat'):
            return obj.isoformat()
        else:
            return str(obj)
    
    def export_to_excel(self, analytics_results: Dict[str, Any],
                       filename: str = "analytics_export.xlsx") -> str:
        """
        Export analytics results to Excel workbook with multiple worksheets
        
        Args:
            analytics_results: Combined analytics results
            filename: Output filename
            
        Returns:
            Path to exported Excel file
        """
        logger.info("📊 Exporting analytics results to Excel workbook")
        
        if not HAS_OPENPYXL:
            logger.warning("⚠️  openpyxl not available, skipping Excel export")
            return ""
        
        try:
            excel_path = self.output_dir / f"{filename.replace('.xlsx', '')}_{self.timestamp}.xlsx"
            
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                
                # Portfolio Summary sheet
                if 'attribution' in analytics_results:
                    portfolio_summary = analytics_results['attribution'].get('portfolio_summary', {})
                    if portfolio_summary:
                        summary_df = pd.DataFrame([portfolio_summary])
                        summary_df.to_excel(writer, sheet_name='Portfolio_Summary', index=False)
                
                # Strategy Contributions sheet
                if 'attribution' in analytics_results:
                    strategy_contrib = analytics_results['attribution'].get('strategy_contribution', {})
                    if strategy_contrib:
                        contrib_data = []
                        for strategy_name, contrib_info in strategy_contrib.items():
                            contrib_data.append({
                                'Strategy': strategy_name,
                                'Total_Contribution': contrib_info.get('total_contribution', 0),
                                'Avg_Daily_Contribution': contrib_info.get('avg_daily_contribution', 0),
                                'Volatility_Contribution': contrib_info.get('volatility_contribution', 0),
                                'Weight': contrib_info.get('weight', 0)
                            })
                        contrib_df = pd.DataFrame(contrib_data)
                        contrib_df.to_excel(writer, sheet_name='Strategy_Contributions', index=False)
                
                # Stop Loss Attribution sheet
                if 'attribution' in analytics_results:
                    sl_attribution = analytics_results['attribution'].get('stop_loss_attribution', {})
                    if sl_attribution:
                        sl_data = []
                        for sl_level, sl_info in sl_attribution.items():
                            sl_data.append({
                                'Stop_Loss_Level': sl_level,
                                'Total_Return': sl_info.get('total_return', 0),
                                'Avg_Daily_Return': sl_info.get('avg_daily_return', 0),
                                'Volatility': sl_info.get('volatility', 0),
                                'Portfolio_Weight': sl_info.get('weight_in_portfolio', 0),
                                'Num_Strategies': sl_info.get('num_strategies', 0)
                            })
                        sl_df = pd.DataFrame(sl_data)
                        sl_df.to_excel(writer, sheet_name='Stop_Loss_Attribution', index=False)
                
                # Take Profit Attribution sheet
                if 'attribution' in analytics_results:
                    tp_attribution = analytics_results['attribution'].get('take_profit_attribution', {})
                    if tp_attribution:
                        tp_data = []
                        for tp_level, tp_info in tp_attribution.items():
                            tp_data.append({
                                'Take_Profit_Level': tp_level,
                                'Total_Return': tp_info.get('total_return', 0),
                                'Avg_Daily_Return': tp_info.get('avg_daily_return', 0),
                                'Volatility': tp_info.get('volatility', 0),
                                'Portfolio_Weight': tp_info.get('weight_in_portfolio', 0),
                                'Num_Strategies': tp_info.get('num_strategies', 0)
                            })
                        tp_df = pd.DataFrame(tp_data)
                        tp_df.to_excel(writer, sheet_name='Take_Profit_Attribution', index=False)
                
                # Risk Metrics sheet  
                if 'risk_metrics' in analytics_results:
                    risk_data = analytics_results['risk_metrics']
                    var_analysis = risk_data.get('var_analysis', {})
                    if var_analysis:
                        risk_summary = {
                            'VaR_95': var_analysis.get('var_95', 0),
                            'VaR_99': var_analysis.get('var_99', 0),
                            'CVaR_95': var_analysis.get('cvar_95', 0),
                            'CVaR_99': var_analysis.get('cvar_99', 0),
                            'Max_Drawdown': risk_data.get('drawdown_analysis', {}).get('max_drawdown', 0),
                            'Portfolio_Volatility': risk_data.get('individual_risk_contributions', {}).get('portfolio_volatility', 0)
                        }
                        risk_df = pd.DataFrame([risk_summary])
                        risk_df.to_excel(writer, sheet_name='Risk_Metrics', index=False)
                
                # Sensitivity Analysis sheet
                if 'sensitivity' in analytics_results:
                    sensitivity_data = analytics_results['sensitivity']
                    
                    # Correlation penalty sensitivity
                    corr_penalty_sens = sensitivity_data.get('correlation_penalty_sensitivity', {})
                    if corr_penalty_sens and 'penalty_results' in corr_penalty_sens:
                        penalty_data = []
                        for key, result in corr_penalty_sens['penalty_results'].items():
                            if 'error' not in result:
                                penalty_data.append({
                                    'Penalty_Weight': result.get('penalty_weight', 0),
                                    'Fitness': result.get('fitness', 0),
                                    'Portfolio_Size': result.get('portfolio_size', 0)
                                })
                        if penalty_data:
                            penalty_df = pd.DataFrame(penalty_data)
                            penalty_df.to_excel(writer, sheet_name='Penalty_Sensitivity', index=False)
                
                # Time Series Data sheet
                if 'attribution' in analytics_results:
                    time_data = analytics_results['attribution'].get('time_attribution', {})
                    if time_data and 'daily_returns' in time_data:
                        time_series_df = pd.DataFrame({
                            'Date': time_data.get('dates', []),
                            'Daily_Return': time_data['daily_returns'],
                            'Cumulative_Return': time_data.get('cumulative_returns', [])
                        })
                        time_series_df.to_excel(writer, sheet_name='Time_Series', index=False)
            
            # Apply formatting to Excel file
            self._format_excel_workbook(excel_path)
            
            logger.info(f"✅ Excel export saved to {excel_path}")
            return str(excel_path)
            
        except Exception as e:
            logger.error(f"❌ Error exporting Excel: {e}")
            return ""
    
    def _format_excel_workbook(self, excel_path: Path):
        """Apply professional formatting to Excel workbook"""
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(excel_path)
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Format each worksheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Format headers
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = border
                
                # Format data cells
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.border = border
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '0.00'
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(excel_path)
            
        except Exception as e:
            logger.warning(f"⚠️  Could not format Excel workbook: {e}")
    
    def export_to_pdf(self, analytics_results: Dict[str, Any],
                     filename: str = "analytics_report.pdf") -> str:
        """
        Export analytics results to professional PDF report
        
        Args:
            analytics_results: Combined analytics results
            filename: Output filename
            
        Returns:
            Path to exported PDF file
        """
        logger.info("📄 Exporting analytics results to PDF report")
        
        if not HAS_REPORTLAB:
            logger.warning("⚠️  reportlab not available, skipping PDF export")
            return ""
        
        try:
            pdf_path = self.output_dir / f"{filename.replace('.pdf', '')}_{self.timestamp}.pdf"
            
            # Create PDF document
            doc = SimpleDocTemplate(str(pdf_path), pagesize=A4,
                                  rightMargin=72, leftMargin=72,
                                  topMargin=72, bottomMargin=18)
            
            # Build PDF content
            story = []
            
            # Title page
            story.append(Paragraph("Advanced Portfolio Analytics Report", self.custom_styles['CustomTitle']))
            story.append(Spacer(1, 12))
            story.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", self.styles['Normal']))
            story.append(Spacer(1, 12))
            story.append(Paragraph("Heavy Optimizer Platform - Production Analysis", self.styles['Normal']))
            story.append(PageBreak())
            
            # Executive Summary
            story.append(Paragraph("Executive Summary", self.custom_styles['CustomHeading']))
            
            if 'attribution' in analytics_results:
                portfolio_summary = analytics_results['attribution'].get('portfolio_summary', {})
                if portfolio_summary:
                    summary_data = [
                        ['Metric', 'Value'],
                        ['Total Return (%)', f"{portfolio_summary.get('total_return', 0):.2f}"],
                        ['Average Daily Return (%)', f"{portfolio_summary.get('avg_daily_return', 0):.4f}"],
                        ['Volatility (%)', f"{portfolio_summary.get('volatility', 0):.2f}"],
                        ['Sharpe Ratio', f"{portfolio_summary.get('sharpe_ratio', 0):.3f}"],
                        ['Maximum Drawdown (%)', f"{portfolio_summary.get('max_drawdown', 0):.2f}"],
                        ['Number of Strategies', f"{portfolio_summary.get('num_strategies', 0)}"],
                        ['Trading Days', f"{portfolio_summary.get('trading_days', 0)}"]
                    ]
                    
                    summary_table = Table(summary_data)
                    summary_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 12),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    
                    story.append(summary_table)
                    story.append(Spacer(1, 12))
            
            # Risk Analysis Section
            story.append(Paragraph("Risk Analysis", self.custom_styles['CustomHeading']))
            
            if 'risk_metrics' in analytics_results:
                risk_data = analytics_results['risk_metrics']
                var_analysis = risk_data.get('var_analysis', {})
                
                if var_analysis:
                    risk_data_table = [
                        ['Risk Metric', 'Value'],
                        ['Value at Risk (95%)', f"{var_analysis.get('var_95', 0):.2f}%"],
                        ['Value at Risk (99%)', f"{var_analysis.get('var_99', 0):.2f}%"],
                        ['Conditional VaR (95%)', f"{var_analysis.get('cvar_95', 0):.2f}%"],
                        ['Conditional VaR (99%)', f"{var_analysis.get('cvar_99', 0):.2f}%"]
                    ]
                    
                    risk_table = Table(risk_data_table)
                    risk_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 12),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    
                    story.append(risk_table)
                    story.append(Spacer(1, 12))
            
            # Top Strategy Contributors
            story.append(Paragraph("Top Strategy Contributors", self.custom_styles['CustomHeading']))
            
            if 'attribution' in analytics_results:
                strategy_contrib = analytics_results['attribution'].get('strategy_contribution', {})
                if strategy_contrib:
                    # Sort by total contribution and take top 10
                    sorted_contribs = sorted(strategy_contrib.items(), 
                                           key=lambda x: x[1].get('total_contribution', 0), 
                                           reverse=True)[:10]
                    
                    contrib_data = [['Strategy Name', 'Total Contribution (%)', 'Weight (%)']]
                    for strategy_name, contrib_info in sorted_contribs:
                        contrib_data.append([
                            strategy_name[:30],  # Truncate long names
                            f"{contrib_info.get('total_contribution', 0):.2f}",
                            f"{contrib_info.get('weight', 0) * 100:.2f}"
                        ])
                    
                    contrib_table = Table(contrib_data)
                    contrib_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('FONTSIZE', (0, 1), (-1, -1), 8)
                    ]))
                    
                    story.append(contrib_table)
            
            # Build PDF
            doc.build(story)
            
            logger.info(f"✅ PDF report saved to {pdf_path}")
            return str(pdf_path)
            
        except Exception as e:
            logger.error(f"❌ Error exporting PDF: {e}")
            return ""
    
    def export_all_formats(self, analytics_results: Dict[str, Any],
                          base_filename: str = "analytics_export") -> Dict[str, Any]:
        """
        Export analytics results to all supported formats
        
        Args:
            analytics_results: Combined analytics results
            base_filename: Base filename for all exports
            
        Returns:
            Dictionary with paths to all exported files
        """
        logger.info("📦 Exporting analytics results to all formats")
        
        export_results = {
            'csv_files': {},
            'json_file': "",
            'excel_file': "",
            'pdf_file': "",
            'export_summary': {}
        }
        
        try:
            # CSV exports
            export_results['csv_files'] = self.export_to_csv(analytics_results, base_filename)
            
            # JSON export
            export_results['json_file'] = self.export_to_json(analytics_results, f"{base_filename}.json")
            
            # Excel export  
            export_results['excel_file'] = self.export_to_excel(analytics_results, f"{base_filename}.xlsx")
            
            # PDF export
            export_results['pdf_file'] = self.export_to_pdf(analytics_results, f"{base_filename}.pdf")
            
            # Create export summary
            export_results['export_summary'] = {
                'export_timestamp': datetime.now().isoformat(),
                'csv_files_count': len(export_results['csv_files']),
                'total_files_created': (
                    len(export_results['csv_files']) + 
                    (1 if export_results['json_file'] else 0) +
                    (1 if export_results['excel_file'] else 0) +
                    (1 if export_results['pdf_file'] else 0)
                ),
                'output_directory': str(self.output_dir),
                'file_sizes': self._get_file_sizes(export_results)
            }
            
            logger.info(f"✅ Exported to {export_results['export_summary']['total_files_created']} files")
            
        except Exception as e:
            logger.error(f"❌ Error in comprehensive export: {e}")
        
        return export_results
    
    def _get_file_sizes(self, export_results: Dict[str, Any]) -> Dict[str, str]:
        """Get file sizes for export summary"""
        file_sizes = {}
        
        try:
            # CSV files
            for csv_type, csv_path in export_results['csv_files'].items():
                if csv_path and Path(csv_path).exists():
                    size_bytes = Path(csv_path).stat().st_size
                    file_sizes[f'csv_{csv_type}'] = f"{size_bytes / 1024:.1f} KB"
            
            # Other files
            for file_type in ['json_file', 'excel_file', 'pdf_file']:
                file_path = export_results.get(file_type, '')
                if file_path and Path(file_path).exists():
                    size_bytes = Path(file_path).stat().st_size
                    file_sizes[file_type] = f"{size_bytes / 1024:.1f} KB"
        
        except Exception as e:
            logger.warning(f"⚠️  Could not calculate file sizes: {e}")
        
        return file_sizes